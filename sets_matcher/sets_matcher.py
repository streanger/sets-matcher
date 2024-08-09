import argparse
import csv
from functools import wraps
import html
import logging
import os
from io import StringIO
from pathlib import Path

import charset_normalizer
import tabulate
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from rich import print
from rich.logging import RichHandler
from rich.table import Table

from sets_matcher.__version__ import __version__

logging.basicConfig(
    level=logging.WARNING,
    format="%(message)s",
    datefmt="[%X]",
    handlers=[RichHandler(markup=True)]
)
logger = logging.getLogger("__name__")


def match_files(files: list[str | Path], max_size: int | None = None) -> tuple[list[str], list[tuple[bool]]]:
    """Returns a list of sets that match in a matrix.

    files: list of files
    max_size: maximu size [B] of single processed file
    """
    if not files:
        raise ValueError("empty list of files")

    pathlib_files = expand_globs(files)

    list_of_sets: list[tuple[str, set[str]]] = []
    for path in pathlib_files:
        try:
            if not max_size is None:
                size = path.stat().st_size
                if size > max_size:
                    logger.warning("file size is too big: [blue]%s[/blue]", path)
                    continue
            content = path.read_bytes()
        except FileNotFoundError as err:
            logger.error("file not found: [blue]%s[/blue]", path)
            continue
        except PermissionError as err:
            logger.error("permission denied: [blue]%s[/blue]", path)
            continue
        except OSError as err:
            logger.error("%s: [blue]%s[/blue]", str(err), path)
            continue
        encoding = charset_normalizer.detect(content)["encoding"]
        if not isinstance(encoding, str):
            logger.error("failed to detect encoding of [blue]%s[blue]", path)
            continue
        items = set(content.decode(encoding).splitlines())
        name = path.stem
        list_of_sets.append((name, items))
        logger.info("[blue]%s[/blue] loaded with %d items", path, len(items))

    # raise error if nothing to process
    if not list_of_sets:
        raise ValueError("nothing to process")

    # create pretty table
    return match_sets(list_of_sets)


def match_sets(list_of_sets: list[set[str]] | list[tuple[str, set[str]]]) -> tuple[list[str], list[tuple[bool]]]:
    """
    Returns a list of sets that match in a matrix.

    accepts list of sets and list of (name, set) tuples
    """
    if not list_of_sets:
        raise ValueError("empty list of sets")

    if all((type(row) is set) for row in list_of_sets):
        header = ["key"] + [str(x) for x in range(1, len(list_of_sets) + 1)]
    elif all(
        (type(row) is tuple and len(row) == 2 and type(row[0]) is str and type(row[1]) is set)
        for row in list_of_sets
    ):
        header = ["key"] + [name for name, _ in list_of_sets]
        list_of_sets = [row for _, row in list_of_sets]
    else:
        raise ValueError("list of sets or list of (name, set) tuples expected")

    keys = sorted(set.union(*list_of_sets))
    table = [[True if word in row else False for word in keys] for row in list_of_sets]
    table.insert(0, keys)
    table = list(zip(*table))  # zip returns iterator of tuples
    return header, table


def parse_args() -> argparse.Namespace:
    """parse cli arguments"""
    if os.name == 'nt':
        os.system('color')
    GREEN = "\u001b[32m"
    YELLOW = "\u001b[33m"
    RED = "\u001b[31m"
    BLUE = "\u001b[34m"
    CYAN = "\u001b[36m"
    RESET = "\u001b[0m"
    version_justed = f"{__version__:<16}"
    box = f"""\
+----------------------------------------------------+
|                    sets-matcher                    |
|----------------------------------------------------|
| version: {version_justed}                          |
|    home: https://github.com/streanger/sets-matcher |
+----------------------------------------------------+"""
    description = f"{GREEN}{box}{RESET}"
    allowed_formats = ("csv", "md", "html", "xlsx")
    parser = argparse.ArgumentParser(description=description, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument("files", nargs="+", type=Path)
    parser.add_argument("--verbose", "-v", action="store_true")
    parser.add_argument("--output", "-o", type=Path)
    parser.add_argument("--format", "-f", choices=allowed_formats)
    parser.add_argument("--index", "-i", action="store_true", help="creates index column")
    args = parser.parse_args()
    if args.verbose:
        logger.setLevel(level=logging.INFO)
    else:
        logger.setLevel(level=logging.WARNING)

    if args.output and not args.format:
        # try to guess format
        allowed_suffixes = [f'.{item}' for item in allowed_formats]
        suffix = Path(args.output).suffix
        if suffix in allowed_suffixes:
            guessed_format = suffix.lstrip('.')
            args.format = guessed_format
            logger.info("format guessed: [blue]%s[/blue]", guessed_format)
            return args
        parser.error(f"--output requires --format or correct suffix {tuple(allowed_suffixes)}")
    elif args.format and not args.output:
        parser.error("--format requires --output")
    elif args.format and args.output:
        # format is stronger then output suffix
        pass
    return args


def main() -> bool | None:
    """main cli entry point"""
    # parse cli arguments
    args = parse_args()

    # match files
    try:
        max_size = 1024 * 1024 * 10  # 10 MB
        header, table = match_files(files=args.files, max_size=max_size)
    except ValueError as err:
        logger.error(err)
        return False

    # write to file
    if args.format == "csv":
        csv = to_csv(header, table, index_column=args.index)
        Path(args.output).write_text(csv, encoding="utf-8")
    elif args.format == "md":
        md = to_markdown(header, table, index_column=args.index)
        Path(args.output).write_text(md, encoding="utf-8")
    elif args.format == "html":
        html = to_html(header, table, index_column=args.index)
        Path(args.output).write_text(html, encoding="utf-8")
    elif args.format == "xlsx":
        to_xlsx(header, table, output=args.output, index_column=args.index)
    else:
        # create pretty table
        rich_table = to_rich_table(header, table, index_column=args.index)
        print(rich_table)
        return None

    logger.info('output saved to: [blue]%s[/blue]', args.output)
    return None


def expand_globs(raw_files: list[str|Path]) -> list[Path]:
    """expand globs in list of files"""
    root_path = Path(".")
    files = []
    for path in raw_files:
        try:
            expanded = list(root_path.glob(str(path)))
        except ValueError as err:
            logger.error("%s", str(err))
            continue
        files.extend(expanded)
    # remove duplicates keep order
    files = list(dict.fromkeys(files))
    return files


def handle_index_column(func):
    @wraps(func)
    def wrapper(header, table, index_column: bool=False, *args, **kwargs):
        if index_column:
            header = list(header)
            header.insert(0, 'Index')
            table = [
                [index] + list(row)
                for index, row in enumerate(table, start=1)
            ]
        return func(header, table, *args, **kwargs)
    return wrapper


@handle_index_column
def to_rich_table(
    header: list[str],
    table: list[list[str | bool]],
    show_lines: bool = False,
    key_style: str = 'green'
) -> Table:
    """convert table with list of lists to rich module pretty table

    Args:
        header (str): table header
        table (list[list]): table body
        show_lines (bool): show table horizontal lines
        key_style (str): key column style - one of: regular, green, blue
    """
    rich_table = Table(
        highlight=True,
        border_style="blue",
        show_lines=show_lines,
        header_style="black on white",
    )
    styles = {
        'regular': "white on black",  # safest option
        'green': "black on green_yellow",  # works fine with no highlight
        'blue': "bold grey3 on deep_sky_blue3",  # better contrast than green
    }
    style = styles.get(key_style, "")
    for index, column in enumerate(header):
        if not index:
            rich_table.add_column(column, style=style, justify="right")
        else:
            rich_table.add_column(column, justify="center")
    for row in table:
        key, *values = row
        key = str(key)
        values = [str(x) for x in values]
        rich_table.add_row(key, *values)
    return rich_table


@handle_index_column
def to_markdown(header: list[str], table: list[list[str | bool]]) -> str:
    """convert table with list of lists to markdown table (github flavored)"""
    def apply_marker(item: str|bool) -> str:
        marker_true = "✓"
        marker_false = ""
        if type(item) is bool:
            if item:
                return marker_true
            else:
                return marker_false
        return item

    table = [[apply_marker(item) for item in row] for row in table]
    md = tabulate.tabulate(table, header, tablefmt="github")
    return md


def to_html(
    header: list[str],
    table: list[list[str | bool]],
    index_column: bool = False,
    title: str = "sets-matcher",
) -> str:
    """convert table with list of lists to html table"""
    if index_column:
        header = list(header)
        header.insert(0, 'Index')
        table = [[index] + list(row) for index, row in enumerate(table, start=1)]

    tab = ' '*4
    table_head = '\n'.join([f"{tab*2}<th><button>{column}</button></th>" for column in header])
    table_head = f'{tab*2}<tr>\n{table_head}\n{tab*2}</tr>'
    table_body = ""
    for row in table:
        cells = []
        for column in row:
            if type(column) is bool:
                if column:
                    column = "✓"
                    cell_class = ' class="marker"'
                else:
                    column = ""
                    cell_class = ""
            else:
                cell_class = ""
            cells.append(f"{tab*3}<td{cell_class}>{column}</td>\n")
        joined_cells = ''.join(cells)
        table_body += f"{tab*2}<tr>\n{joined_cells}{tab*2}</tr>\n"
    table_body = table_body.rstrip()

    # TODO: read style & script from files
    style = """\
    .styled-table {
        border-collapse: collapse;
        margin-left: auto;
        margin-right: auto;
        font-size: 0.9em;
        font-family: sans-serif;
        min-width: 400px;
    }
    .styled-table thead tr {
        background-color: #009879;
        color: #ffffff;
        position: sticky;
        top: 0;
    }
    .styled-table td {
        padding: 12px 15px;
        text-align: center;
    }
    .styled-table td:first-child {
        padding: 12px 15px;
        text-align: right;
    }
    .styled-table tbody tr {
        border-bottom: 1px solid #dddddd;
    }
    .styled-table th {
        padding: 0;
        text-align: center;
    }
    .styled-table th button {
        background-color: transparent;
        border: none;
        font: inherit;
        color: inherit;
        height: 100%;
        width: 100%;
        padding: 12px 15px;
        display: inline-block;
    }
    .styled-table th button::after {
        content: "\\00a0\\00a0";
        font-family: 'Courier New', Courier, monospace
    }
    .styled-table th button[direction="ascending"]::after {
        content: "\\00a0▲";
    }
    .styled-table th button[direction="descending"]::after {
        content: "\\00a0▼";
    }
    .marker {
    background-color: #cccccc;
    border-radius: 10px;
    }"""

    index_column_script = """\
function main() {
    var table = document.getElementsByTagName("table")[0];
    var header = table.getElementsByTagName("tr")[0];
    var headers = header.getElementsByTagName("th");
    for (var i = 1; i < headers.length; i++) {
        var btn = headers[i].getElementsByTagName("button")[0];
        btn.setAttribute("onclick", `table_sorter(${i})`);
    }
}

function table_sorter(column) {
    var table = document.getElementsByTagName("table")[0];
    var tableBody = table.getElementsByTagName("tbody")[0];
    var columnButton = table.getElementsByTagName("tr")[0].getElementsByTagName("th")[column].getElementsByTagName("button")[0];
    var direction = columnButton.getAttribute("direction");
    if (direction == "ascending") {
        direction = "descending";
    } else {
        direction = "ascending";
    }
    var rows = Array.from(table.getElementsByTagName("tr")).slice(1);
    rows.sort(function(a, b) {
        var x = a.cells[column].textContent.toLowerCase();
        var y = b.cells[column].textContent.toLowerCase();
        if (direction === "ascending") {
            return x - y || x.localeCompare(y);
        } else {
            return y - x || y.localeCompare(x);
        }
    });
    tableBody.innerHTML = '';
    rows.forEach((row, i) => {
        var newRow = row.cloneNode(true);
        newRow.cells[0].textContent = i + 1;
        tableBody.appendChild(newRow);
    });

    // show direction using arrow icon
    var header = table.getElementsByTagName("tr")[0];
    var headers = header.getElementsByTagName("th");
    for (var i = 0; i < headers.length; i++) {
        var btn = headers[i].getElementsByTagName("button")[0];
        if (i == column) {
            btn.setAttribute("direction", direction);
        } else {
            btn.setAttribute("direction", "");
        }
    }
}"""

    script = """\
function main() {
    var table = document.getElementsByTagName("table")[0];
    var header = table.getElementsByTagName("tr")[0];
    var headers = header.getElementsByTagName("th");
    for (var i = 0; i < headers.length; i++) {
        var btn = headers[i].getElementsByTagName("button")[0];
        btn.setAttribute("onclick", `table_sorter(${i})`);
    }
}

function table_sorter(column) {
    var table = document.getElementsByTagName("table")[0];
    var tableBody = table.getElementsByTagName("tbody")[0];
    var columnButton = table.getElementsByTagName("tr")[0].getElementsByTagName("th")[column].getElementsByTagName("button")[0];
    var direction = columnButton.getAttribute("direction");
    if (direction == "ascending") {
        direction = "descending";
    } else {
        direction = "ascending";
    }
    var rows = Array.from(table.getElementsByTagName("tr")).slice(1);
    rows.sort(function(a, b) {
        var x = a.getElementsByTagName("td")[column].textContent.toLowerCase();
        var y = b.getElementsByTagName("td")[column].textContent.toLowerCase();
        if (direction === "ascending") {
            return x.localeCompare(y);
        } else {
            return y.localeCompare(x);
        }
    });
    rows.forEach(function(row) {
        tableBody.appendChild(row);
    });

    // show direction using arrow icon
    var header = table.getElementsByTagName("tr")[0];
    var headers = header.getElementsByTagName("th");
    for (var i = 0; i < headers.length; i++) {
        var btn = headers[i].getElementsByTagName("button")[0];
        if (i == column) {
            btn.setAttribute("direction", direction);
        } else {
            btn.setAttribute("direction", "");
        }
    }
}"""
    # dirty way to keep old style & previous performance
    if index_column:
        script = index_column_script

    title = html.escape(title)
    template = f"""\
<!DOCTYPE html>
<html>
<head>
    <title>{title}</title>
    <meta charset="utf-8">
    <style>
{style}
    </style>
    <script>
{script}
    </script>
</head>
<body onload=main()>
    <table class="styled-table">
    <thead>
{table_head}
    </thead>
    <tbody>
{table_body}
    </tbody>
    </table>
</body>
</html>\
"""
    return template


@handle_index_column
def to_csv(header: list[str], table: list[list[str | bool]]) -> str:
    """convert table with list of lists to csv table"""
    output = StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(header)
    for row in table:
        writer.writerow(row)
    csv_string = output.getvalue()
    return csv_string


@handle_index_column
def to_xlsx(
    header: list[str],
    table: list[list[str | bool]],
    output: str | Path = "out.xlsx"
) -> None:
    """convert table with list of lists to xlsx table"""
    wb = Workbook()
    ws = wb.active

    fill = PatternFill(start_color="009879", end_color="009879", fill_type="solid")
    checkmark_fill = PatternFill(start_color="dddddd", end_color="dddddd", fill_type="solid")
    align = Alignment(horizontal="center")

    for col_num, value in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num, value=value)
        cell.fill = fill
        cell.alignment = align
        ws.column_dimensions[cell.column_letter].width = max(len(str(value)), 4)

    for index, row in enumerate(table, 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=index, column=col_num)
            if isinstance(value, bool):
                cell.value = '✔' if value else ''
                cell.alignment = align
                if value:
                    cell.fill = checkmark_fill
            else:
                cell.value = value
                ws.column_dimensions[cell.column_letter].width = max(ws.column_dimensions[cell.column_letter].width, len(str(value)))
    if type(output) is Path:
        output = str(output)
    wb.save(output)


if __name__ == "__main__":
    main()
