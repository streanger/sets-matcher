# https://docs.pytest.org/en/7.1.x/explanation/goodpractices.html#tests-as-part-of-application-code
import hashlib
from pathlib import Path
from openpyxl import load_workbook
from sets_matcher import match_sets, to_csv, to_html, to_markdown, to_xlsx

SETS_EXAMPLE = [
    {'some', 'thing', 'here'},
    {'this', 'is', 'sparta', 'some', 'thing', 'here'},
    {'now', 'some', 'this', 'thing', 'here'},
]

def test_export_html():
    """
    to create & open test.html:
        Path('test.html').write_text(html, encoding='utf-8')
        webbrowser.open('test.html')
    """
    header, table = match_sets(SETS_EXAMPLE)
    html = to_html(header, table, index_column=False)
    test_html = Path('test/out/test.html').read_text(encoding='utf-8')
    assert html == test_html


def test_export_html_index_column():
    """
    to create & open test-index-column.html:
        Path('test-index-column.html').write_text(html, encoding='utf-8')
        webbrowser.open('test-index-column.html')
    """
    header, table = match_sets(SETS_EXAMPLE)
    # call twice for idempotence testing
    html = to_html(header, table, index_column=True)
    html = to_html(header, table, index_column=True)
    test_html = Path('test/out/test-index-column.html').read_text(encoding='utf-8')
    assert html == test_html


def test_export_csv():
    """
    to create test.csv:
        Path('test.csv').write_text(csv, encoding='utf-8')
    """
    header, table = match_sets(SETS_EXAMPLE)
    # call twice for idempotence testing
    csv = to_csv(header, table)
    csv = to_csv(header, table)
    test_csv = Path('test/out/test.csv').read_text(encoding='utf-8')
    assert csv == test_csv


def test_export_markdown():
    """
    to create test.md:
        Path('test.md').write_text(md, encoding='utf-8')
    """
    header, table = match_sets(SETS_EXAMPLE)
    md = to_markdown(header, table)
    test_md = Path('test/out/test.md').read_text(encoding='utf-8')
    assert md == test_md


def read_xlsx(file_path):
    """read xlsx content for test purposes"""
    wb = load_workbook(filename=file_path)
    sheet = wb.active
    return [(row[0].value, row[1].value) for row in sheet.iter_rows()]


def test_export_xlsx():
    """test export to xlsx file"""
    # remove possible file
    new_xlsx_path = Path('test/out/new.xlsx')
    new_xlsx_path.unlink(missing_ok=True)

    # create xlsx file
    header, table = match_sets(SETS_EXAMPLE)
    to_xlsx(header, table, output=new_xlsx_path)

    # compare workbooks content
    test_xlsx_path = Path('test/out/test.xlsx')
    test_wb = read_xlsx(test_xlsx_path)
    new_wb = read_xlsx(new_xlsx_path)
    assert test_wb == new_wb

    # make sure new file is removed after test
    new_xlsx_path.unlink(missing_ok=True)
    assert not new_xlsx_path.exists()


def test_input_data():
    header, table = match_sets(SETS_EXAMPLE)

    # different types of data
    # INFO: it should be the case for static typin,
    # but as long as its not fully correct
    # its better to test that way, than nothing
    header = tuple(header)
    table = [list(row) for row in table]

    html = to_html(header, table, index_column=True)
    csv = to_csv(header, table, index_column=True)
    md = to_markdown(header, table, index_column=True)

    new_xlsx_path = Path('test/out/new.xlsx')
    new_xlsx_path.unlink(missing_ok=True)
    to_xlsx(header, table, index_column=True, output=new_xlsx_path)
    new_xlsx_path.unlink(missing_ok=True)
    assert not new_xlsx_path.exists()
